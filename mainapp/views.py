from django.shortcuts import render, redirect
from django.views import generic
import requests
import socket
from openpyxl import load_workbook

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ObjectDoesNotExist
from rest_framework import status

from .models import Website
from .serializers import WebsiteSerializer

''' Home page view '''
class IndexView(generic.TemplateView):
    template_name = 'index.html'

    def get(self, request, *args, **kwargs):
        self.extra_context = {
            'websites': Website.objects.all()
        }
        return super().get(request, *args, **kwargs)

"""
    Проверка сайта.
    Открывает websites.xlsx (где хранится url адреса)
    и пробегает по индексам от count_websites до длины excel файла. 
    count_websites - количество проверенных сайтов.
    Это сделано для того чтобы продолжить проверку с места срыва    
"""
def start_check(request):
    file = 'mediafiles/websites.xlsx'
    wb = load_workbook(file)
    sheet_name = wb.sheetnames[0]
    sheet = wb.get_sheet_by_name(sheet_name)
    data = list(sheet.values)
    count_websites = Website.objects.all().count()
    for i in range(count_websites+1, len(data)):
        url = data[i][0]
        if '/' in url:
            url = str(url).split('/')[0]
        """Если сайт не существует, то load_time, ip_addressm http_code будет равно None"""
        try:
            response = requests.get("http://"+url)
            load_time = response.elapsed.total_seconds()
            ip_address = socket.gethostbyname(url)
            http_code = response.status_code
        except requests.ConnectionError as exception:
            load_time,ip_address,http_code = 'None','None', 'None'
        Website.objects.create(url=url, ip_address=ip_address, load_time=load_time, http_code=http_code)
    return redirect('/')


"""
    Вывод результата по url
"""
@api_view(["GET"])
@csrf_exempt
@permission_classes([IsAuthenticated])
def get_website(request):
    try:
        url = request.GET['url']
        website = Website.objects.get(url=url)
        serializer = WebsiteSerializer(website)
        return JsonResponse({'website': serializer.data}, safe=False, status=status.HTTP_200_OK)
    except ObjectDoesNotExist as e:
        return JsonResponse({'error': str(e)}, safe=False, status=status.HTTP_404_NOT_FOUND)
    except Exception:
        return JsonResponse({'error': 'Something terrible went wrong'}, safe=False,
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)



